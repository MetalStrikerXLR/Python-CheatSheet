# use PyPI to search for packages. Use terminal to install packages using pip
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_excel_sheet(filename: object) -> object:
    # Load excel file in wb variable
    assert isinstance(filename, object)
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # identify spreadsheet max rows
    max_rows = sheet.max_row

    # access 3rd column of rows 2-4, correct prices and add them in 4th column
    for row in range(2, max_rows + 1):
        cell = sheet.cell(row, 3)
        corr_values = cell.value * 0.9
        corr_value_cell = sheet.cell(row, 4)
        corr_value_cell.value = corr_values

    # Provide reference of values to be used in BarChart
    values = Reference(sheet, min_row=2, max_row=max_rows, min_col=4, max_col=4)

    # Create a BarChart
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')

    # save excel file
    wb.save(filename)
